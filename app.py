from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from database import get_db, init_db, DATABASE_URL
from config import TALLER
import io, os

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'taller_motos_secreto_2024')

# ─── DB HELPERS ─────────────────────────────────────────────────────────────

def _sql(sql):
    return sql.replace('?', '%s') if DATABASE_URL else sql

def fetchall(db, sql, params=()):
    c = db.cursor()
    c.execute(_sql(sql), params)
    rows = c.fetchall()
    if DATABASE_URL:
        return list(rows)
    return [dict(r) for r in rows]

def fetchone(db, sql, params=()):
    c = db.cursor()
    c.execute(_sql(sql), params)
    row = c.fetchone()
    if row is None: return None
    return dict(row) if not DATABASE_URL else row

def execute(db, sql, params=()):
    c = db.cursor()
    c.execute(_sql(sql), params)
    return c

def scalar(db, sql, params=()):
    c = db.cursor()
    c.execute(_sql(sql), params)
    row = c.fetchone()
    if row is None: return 0
    return list(row.values())[0] if DATABASE_URL else row[0]

# ─── INIT ────────────────────────────────────────────────────────────────────

@app.before_request
def setup():
    if not hasattr(app, '_db_initialized'):
        init_db()
        app._db_initialized = True

# ─── INDEX ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    db = get_db()
    stats = {
        'clientes':        scalar(db, 'SELECT COUNT(*) FROM clientes'),
        'motos':           scalar(db, 'SELECT COUNT(*) FROM vehiculos'),
        'ordenes_abiertas': scalar(db, "SELECT COUNT(*) FROM ordenes WHERE estado='abierta'"),
        'ordenes_cerradas': scalar(db, "SELECT COUNT(*) FROM ordenes WHERE estado='cerrada'"),
    }
    ordenes_recientes = fetchall(db, '''
        SELECT o.id, o.fecha_ingreso, o.estado,
               c.nombre || ' ' || c.apellido AS cliente,
               v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS moto
        FROM ordenes o
        JOIN clientes c ON o.cliente_id = c.id
        JOIN vehiculos v ON o.vehiculo_id = v.id
        ORDER BY o.created_at DESC LIMIT 5
    ''')
    db.close()
    return render_template('index.html', stats=stats, ordenes_recientes=ordenes_recientes)

# ─── CLIENTES ────────────────────────────────────────────────────────────────

@app.route('/clientes')
def clientes():
    q = request.args.get('q', '')
    db = get_db()
    if q:
        rows = fetchall(db,
            "SELECT * FROM clientes WHERE nombre LIKE ? OR apellido LIKE ? OR telefono LIKE ? ORDER BY apellido",
            (f'%{q}%', f'%{q}%', f'%{q}%'))
    else:
        rows = fetchall(db, 'SELECT * FROM clientes ORDER BY apellido')
    db.close()
    return render_template('clientes.html', clientes=rows, q=q)

@app.route('/clientes/nuevo', methods=['GET', 'POST'])
def nuevo_cliente():
    if request.method == 'POST':
        nombre   = request.form['nombre'].strip()
        apellido = request.form['apellido'].strip()
        direccion = request.form.get('direccion','').strip()
        telefono  = request.form.get('telefono','').strip()
        email     = request.form.get('email','').strip()
        dni       = request.form.get('dni','').strip()
        if not nombre or not apellido:
            flash('Nombre y apellido son obligatorios.', 'danger')
        else:
            db = get_db()
            execute(db, 'INSERT INTO clientes (nombre,apellido,direccion,telefono,email,dni) VALUES (?,?,?,?,?,?)',
                    (nombre, apellido, direccion, telefono, email, dni))
            db.commit(); db.close()
            flash('Cliente agregado.', 'success')
            return redirect(url_for('clientes'))
    return render_template('cliente_form.html', cliente=None)

@app.route('/clientes/<int:id>/editar', methods=['GET', 'POST'])
def editar_cliente(id):
    db = get_db()
    cliente = fetchone(db, 'SELECT * FROM clientes WHERE id=?', (id,))
    if not cliente:
        db.close(); flash('Cliente no encontrado.', 'danger')
        return redirect(url_for('clientes'))
    if request.method == 'POST':
        execute(db, 'UPDATE clientes SET nombre=?,apellido=?,direccion=?,telefono=?,email=?,dni=? WHERE id=?',
                (request.form['nombre'].strip(), request.form['apellido'].strip(),
                 request.form.get('direccion','').strip(), request.form.get('telefono','').strip(),
                 request.form.get('email','').strip(), request.form.get('dni','').strip(), id))
        db.commit(); db.close()
        flash('Cliente actualizado.', 'success')
        return redirect(url_for('clientes'))
    db.close()
    return render_template('cliente_form.html', cliente=cliente)

@app.route('/clientes/<int:id>')
def ver_cliente(id):
    db = get_db()
    cliente  = fetchone(db, 'SELECT * FROM clientes WHERE id=?', (id,))
    motos    = fetchall(db, 'SELECT * FROM vehiculos WHERE cliente_id=?', (id,))
    ordenes  = fetchall(db, '''
        SELECT o.*, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS moto
        FROM ordenes o JOIN vehiculos v ON o.vehiculo_id=v.id
        WHERE o.cliente_id=? ORDER BY o.fecha_ingreso DESC
    ''', (id,))
    db.close()
    return render_template('cliente_detalle.html', cliente=cliente, motos=motos, ordenes=ordenes)

@app.route('/clientes/<int:id>/eliminar', methods=['POST'])
def eliminar_cliente(id):
    db = get_db()
    execute(db, 'DELETE FROM clientes WHERE id=?', (id,))
    db.commit(); db.close()
    flash('Cliente eliminado.', 'success')
    return redirect(url_for('clientes'))

# ─── MOTOS ───────────────────────────────────────────────────────────────────

@app.route('/motos')
def motos():
    q = request.args.get('q', '')
    db = get_db()
    if q:
        rows = fetchall(db, '''
            SELECT v.*, c.nombre || ' ' || c.apellido AS cliente
            FROM vehiculos v LEFT JOIN clientes c ON v.cliente_id=c.id
            WHERE v.marca LIKE ? OR v.modelo LIKE ? OR v.patente LIKE ?
               OR v.cilindrada LIKE ? OR v.tipo_moto LIKE ?
            ORDER BY v.marca
        ''', (f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%'))
    else:
        rows = fetchall(db, '''
            SELECT v.*, c.nombre || ' ' || c.apellido AS cliente
            FROM vehiculos v LEFT JOIN clientes c ON v.cliente_id=c.id ORDER BY v.marca
        ''')
    db.close()
    return render_template('motos.html', motos=rows, q=q)

@app.route('/motos/nueva', methods=['GET', 'POST'])
def nueva_moto():
    db = get_db()
    clientes_list = fetchall(db, "SELECT id, nombre || ' ' || apellido AS nombre FROM clientes ORDER BY apellido")
    if request.method == 'POST':
        marca      = request.form['marca'].strip()
        modelo     = request.form['modelo'].strip()
        patente    = request.form['patente'].strip().upper()
        anio       = request.form.get('anio','').strip() or None
        motor      = request.form.get('motor','').strip() or None
        vin        = request.form.get('vin','').strip() or None
        color      = request.form.get('color','').strip() or None
        combustible = request.form.get('combustible','').strip() or None
        cilindrada  = request.form.get('cilindrada','').strip() or None
        tipo_moto   = request.form.get('tipo_moto','').strip() or None
        cliente_id  = request.form.get('cliente_id') or None
        if not marca or not modelo or not patente:
            flash('Marca, modelo y dominio son obligatorios.', 'danger')
        else:
            try:
                execute(db, '''INSERT INTO vehiculos
                    (marca,modelo,patente,anio,motor,vin,color,combustible,cilindrada,tipo_moto,cliente_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                        (marca, modelo, patente, anio, motor, vin, color, combustible,
                         cilindrada, tipo_moto, cliente_id))
                db.commit(); db.close()
                flash('Moto agregada.', 'success')
                return redirect(url_for('motos'))
            except Exception:
                flash('El dominio ya existe.', 'danger')
    db.close()
    return render_template('moto_form.html', moto=None, clientes=clientes_list)

@app.route('/motos/<int:id>/editar', methods=['GET', 'POST'])
def editar_moto(id):
    db = get_db()
    moto = fetchone(db, 'SELECT * FROM vehiculos WHERE id=?', (id,))
    clientes_list = fetchall(db, "SELECT id, nombre || ' ' || apellido AS nombre FROM clientes ORDER BY apellido")
    if request.method == 'POST':
        try:
            execute(db, '''UPDATE vehiculos SET
                marca=?,modelo=?,patente=?,anio=?,motor=?,vin=?,color=?,
                combustible=?,cilindrada=?,tipo_moto=?,cliente_id=? WHERE id=?''',
                    (request.form['marca'].strip(), request.form['modelo'].strip(),
                     request.form['patente'].strip().upper(),
                     request.form.get('anio','').strip() or None,
                     request.form.get('motor','').strip() or None,
                     request.form.get('vin','').strip() or None,
                     request.form.get('color','').strip() or None,
                     request.form.get('combustible','').strip() or None,
                     request.form.get('cilindrada','').strip() or None,
                     request.form.get('tipo_moto','').strip() or None,
                     request.form.get('cliente_id') or None, id))
            db.commit(); db.close()
            flash('Moto actualizada.', 'success')
            return redirect(url_for('motos'))
        except Exception:
            flash('El dominio ya existe.', 'danger')
    db.close()
    return render_template('moto_form.html', moto=moto, clientes=clientes_list)

@app.route('/motos/<int:id>/eliminar', methods=['POST'])
def eliminar_moto(id):
    db = get_db()
    execute(db, 'DELETE FROM vehiculos WHERE id=?', (id,))
    db.commit(); db.close()
    flash('Moto eliminada.', 'success')
    return redirect(url_for('motos'))

# ─── ÓRDENES ─────────────────────────────────────────────────────────────────

@app.route('/ordenes')
def ordenes():
    q = request.args.get('q', '')
    estado = request.args.get('estado', '')
    db = get_db()
    sql = '''
        SELECT o.id, o.fecha_ingreso, o.fecha_egreso, o.estado,
               c.nombre || ' ' || c.apellido AS cliente,
               v.marca || ' ' || v.modelo AS moto, v.patente
        FROM ordenes o
        JOIN clientes c ON o.cliente_id=c.id
        JOIN vehiculos v ON o.vehiculo_id=v.id WHERE 1=1
    '''
    params = []
    if q:
        sql += ' AND (c.nombre LIKE ? OR c.apellido LIKE ? OR v.patente LIKE ?)'
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if estado:
        sql += ' AND o.estado=?'
        params.append(estado)
    sql += ' ORDER BY o.created_at DESC'
    rows = fetchall(db, sql, params)
    db.close()
    return render_template('ordenes.html', ordenes=rows, q=q, estado=estado)

def _orden_query():
    return '''
        SELECT o.*,
               c.nombre || ' ' || c.apellido AS cliente_nombre,
               c.apellido AS cliente_apellido, c.nombre AS cliente_nombre2,
               c.telefono AS cliente_tel, c.direccion AS cliente_dir,
               c.email AS cliente_email, c.dni AS cliente_dni,
               v.marca, v.modelo, v.patente, v.anio, v.motor,
               v.vin, v.color, v.combustible, v.cilindrada, v.tipo_moto
        FROM ordenes o
        JOIN clientes c ON o.cliente_id=c.id
        JOIN vehiculos v ON o.vehiculo_id=v.id
        WHERE o.id=?
    '''

def _save_repuestos(db, orden_id, form):
    execute(db, 'DELETE FROM orden_repuestos WHERE orden_id=?', (orden_id,))
    descs = form.getlist('rep_desc')
    cants = form.getlist('rep_cantidad')
    for d, c in zip(descs, cants):
        if d.strip():
            execute(db, 'INSERT INTO orden_repuestos (orden_id, descripcion, cantidad) VALUES (?,?,?)',
                    (orden_id, d.strip(), float(c or 1)))

@app.route('/ordenes/nueva', methods=['GET', 'POST'])
def nueva_orden():
    db = get_db()
    clientes_list = fetchall(db, "SELECT id, nombre || ' ' || apellido AS nombre FROM clientes ORDER BY apellido")
    motos_list = fetchall(db, '''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''')
    if request.method == 'POST':
        execute(db, '''INSERT INTO ordenes
            (vehiculo_id,cliente_id,fecha_ingreso,fecha_estimada,fecha_egreso,kilometros,
             receptor_servicio,descripcion_trabajo,diagnostico,repuestos,observaciones,estado)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
            (request.form['vehiculo_id'], request.form['cliente_id'],
             request.form['fecha_ingreso'],
             request.form.get('fecha_estimada') or None,
             request.form.get('fecha_egreso') or None,
             request.form.get('kilometros') or None,
             request.form.get('receptor_servicio','').strip(),
             request.form.get('descripcion_trabajo','').strip(),
             request.form.get('diagnostico','').strip(),
             request.form.get('repuestos','').strip(),
             request.form.get('observaciones','').strip(),
             request.form.get('estado','abierta')))
        if DATABASE_URL:
            oid = scalar(db, 'SELECT lastval()')
        else:
            oid = db.execute('SELECT last_insert_rowid()').fetchone()[0]
        _save_repuestos(db, oid, request.form)
        db.commit(); db.close()
        flash('Orden creada.', 'success')
        return redirect(url_for('ordenes'))
    db.close()
    return render_template('orden_form.html', orden=None, clientes=clientes_list, motos=motos_list, repuestos=[])

@app.route('/ordenes/<int:id>')
def ver_orden(id):
    db = get_db()
    orden = fetchone(db, _orden_query(), (id,))
    repuestos = fetchall(db, 'SELECT * FROM orden_repuestos WHERE orden_id=? ORDER BY id', (id,))
    db.close()
    return render_template('orden_detalle.html', orden=orden, repuestos=repuestos)

@app.route('/ordenes/<int:id>/imprimir')
def imprimir_orden(id):
    db = get_db()
    orden = fetchone(db, _orden_query(), (id,))
    repuestos = fetchall(db, 'SELECT * FROM orden_repuestos WHERE orden_id=? ORDER BY id', (id,))
    db.close()
    return render_template('orden_print.html', orden=orden, repuestos=repuestos, taller=TALLER)

@app.route('/ordenes/<int:id>/editar', methods=['GET', 'POST'])
def editar_orden(id):
    db = get_db()
    orden = fetchone(db, 'SELECT * FROM ordenes WHERE id=?', (id,))
    clientes_list = fetchall(db, "SELECT id, nombre || ' ' || apellido AS nombre FROM clientes ORDER BY apellido")
    motos_list = fetchall(db, '''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''')
    if request.method == 'POST':
        execute(db, '''UPDATE ordenes SET
            vehiculo_id=?,cliente_id=?,fecha_ingreso=?,fecha_estimada=?,fecha_egreso=?,
            kilometros=?,receptor_servicio=?,descripcion_trabajo=?,diagnostico=?,
            repuestos=?,observaciones=?,estado=? WHERE id=?''',
            (request.form['vehiculo_id'], request.form['cliente_id'],
             request.form['fecha_ingreso'],
             request.form.get('fecha_estimada') or None,
             request.form.get('fecha_egreso') or None,
             request.form.get('kilometros') or None,
             request.form.get('receptor_servicio','').strip(),
             request.form.get('descripcion_trabajo','').strip(),
             request.form.get('diagnostico','').strip(),
             request.form.get('repuestos','').strip(),
             request.form.get('observaciones','').strip(),
             request.form.get('estado','abierta'), id))
        _save_repuestos(db, id, request.form)
        db.commit(); db.close()
        flash('Orden actualizada.', 'success')
        return redirect(url_for('ver_orden', id=id))
    repuestos = fetchall(db, 'SELECT * FROM orden_repuestos WHERE orden_id=? ORDER BY id', (id,))
    db.close()
    return render_template('orden_form.html', orden=orden, clientes=clientes_list, motos=motos_list, repuestos=repuestos)

@app.route('/ordenes/<int:id>/cerrar', methods=['POST'])
def cerrar_orden(id):
    from datetime import date
    db = get_db()
    execute(db, "UPDATE ordenes SET estado='cerrada', fecha_egreso=? WHERE id=?",
            (date.today().isoformat(), id))
    db.commit(); db.close()
    flash('Orden cerrada.', 'success')
    return redirect(url_for('ver_orden', id=id))

@app.route('/ordenes/<int:id>/eliminar', methods=['POST'])
def eliminar_orden(id):
    db = get_db()
    execute(db, 'DELETE FROM ordenes WHERE id=?', (id,))
    db.commit(); db.close()
    flash('Orden eliminada.', 'success')
    return redirect(url_for('ordenes'))

# ─── PRESUPUESTOS ────────────────────────────────────────────────────────────

@app.route('/presupuestos')
def presupuestos():
    q = request.args.get('q', '')
    estado = request.args.get('estado', '')
    db = get_db()
    sql = '''
        SELECT p.id, p.fecha, p.estado, p.descripcion,
               c.nombre || ' ' || c.apellido AS cliente,
               v.marca || ' ' || v.modelo AS moto, v.patente,
               COALESCE((SELECT SUM(cantidad * precio_unitario) FROM presupuesto_items WHERE presupuesto_id=p.id), 0) AS total
        FROM presupuestos p
        JOIN clientes c ON p.cliente_id=c.id
        LEFT JOIN vehiculos v ON p.vehiculo_id=v.id WHERE 1=1
    '''
    params = []
    if q:
        sql += ' AND (c.nombre LIKE ? OR c.apellido LIKE ? OR v.patente LIKE ?)'
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if estado:
        sql += ' AND p.estado=?'
        params.append(estado)
    sql += ' ORDER BY p.created_at DESC'
    rows = fetchall(db, sql, params)
    db.close()
    return render_template('presupuestos.html', presupuestos=rows, q=q, estado=estado)

def _presup_query():
    return '''
        SELECT p.*,
               c.nombre || ' ' || c.apellido AS cliente_nombre,
               c.telefono AS cliente_tel, c.direccion AS cliente_dir,
               v.marca, v.modelo, v.patente, v.anio, v.cilindrada, v.tipo_moto
        FROM presupuestos p
        JOIN clientes c ON p.cliente_id=c.id
        LEFT JOIN vehiculos v ON p.vehiculo_id=v.id
        WHERE p.id=?
    '''

def _save_items(db, pid, form):
    tipos    = form.getlist('item_tipo')
    descs    = form.getlist('item_desc')
    cantidades = form.getlist('item_cantidad')
    precios  = form.getlist('item_precio')
    stocks   = form.getlist('item_stock')
    for i, (t, d, c, p) in enumerate(zip(tipos, descs, cantidades, precios)):
        if d.strip():
            s = 1 if i < len(stocks) and stocks[i] == '1' else 0
            execute(db, 'INSERT INTO presupuesto_items (presupuesto_id,tipo,descripcion,cantidad,precio_unitario,en_stock) VALUES (?,?,?,?,?,?)',
                    (pid, t, d.strip(), float(c or 1), float(p or 0), s))

@app.route('/presupuestos/nuevo', methods=['GET', 'POST'])
def nuevo_presupuesto():
    db = get_db()
    clientes_list = fetchall(db, "SELECT id, nombre || ' ' || apellido AS nombre FROM clientes ORDER BY apellido")
    motos_list = fetchall(db, '''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''')
    if request.method == 'POST':
        sql_insert = '''INSERT INTO presupuestos
            (cliente_id,vehiculo_id,fecha,valido_hasta,descripcion,observaciones,atendido_por,condicion_venta,estado)
            VALUES (?,?,?,?,?,?,?,?,?)'''
        if DATABASE_URL:
            sql_insert += ' RETURNING id'
        cur = execute(db, sql_insert,
            (request.form['cliente_id'],
             request.form.get('vehiculo_id') or None,
             request.form['fecha'],
             request.form.get('valido_hasta') or None,
             request.form.get('descripcion','').strip(),
             request.form.get('observaciones','').strip(),
             request.form.get('atendido_por','').strip(),
             request.form.get('condicion_venta','').strip(),
             request.form.get('estado','pendiente')))
        if DATABASE_URL:
            pid = cur.fetchone()['id']
        else:
            pid = cur.lastrowid
        _save_items(db, pid, request.form)
        db.commit(); db.close()
        flash('Presupuesto creado.', 'success')
        return redirect(url_for('ver_presupuesto', id=pid))
    db.close()
    return render_template('presupuesto_form.html', presupuesto=None, clientes=clientes_list, motos=motos_list, items=[])

@app.route('/presupuestos/<int:id>')
def ver_presupuesto(id):
    db = get_db()
    p     = fetchone(db, _presup_query(), (id,))
    items = fetchall(db, 'SELECT * FROM presupuesto_items WHERE presupuesto_id=? ORDER BY tipo, id', (id,))
    total = sum(i['cantidad'] * i['precio_unitario'] for i in items)
    db.close()
    return render_template('presupuesto_detalle.html', p=p, items=items, total=total)

@app.route('/presupuestos/<int:id>/editar', methods=['GET', 'POST'])
def editar_presupuesto(id):
    db = get_db()
    presupuesto = fetchone(db, 'SELECT * FROM presupuestos WHERE id=?', (id,))
    clientes_list = fetchall(db, "SELECT id, nombre || ' ' || apellido AS nombre FROM clientes ORDER BY apellido")
    motos_list = fetchall(db, '''
        SELECT v.id, v.marca || ' ' || v.modelo || ' (' || v.patente || ')' AS descripcion, v.cliente_id
        FROM vehiculos v ORDER BY v.marca
    ''')
    items = fetchall(db, 'SELECT * FROM presupuesto_items WHERE presupuesto_id=? ORDER BY tipo, id', (id,))
    if request.method == 'POST':
        execute(db, '''UPDATE presupuestos SET
            cliente_id=?,vehiculo_id=?,fecha=?,valido_hasta=?,descripcion=?,
            observaciones=?,atendido_por=?,condicion_venta=?,estado=? WHERE id=?''',
            (request.form['cliente_id'],
             request.form.get('vehiculo_id') or None,
             request.form['fecha'],
             request.form.get('valido_hasta') or None,
             request.form.get('descripcion','').strip(),
             request.form.get('observaciones','').strip(),
             request.form.get('atendido_por','').strip(),
             request.form.get('condicion_venta','').strip(),
             request.form.get('estado','pendiente'), id))
        execute(db, 'DELETE FROM presupuesto_items WHERE presupuesto_id=?', (id,))
        _save_items(db, id, request.form)
        db.commit(); db.close()
        flash('Presupuesto actualizado.', 'success')
        return redirect(url_for('ver_presupuesto', id=id))
    db.close()
    return render_template('presupuesto_form.html', presupuesto=presupuesto, clientes=clientes_list, motos=motos_list, items=items)

@app.route('/presupuestos/<int:id>/estado/<nuevo_estado>', methods=['POST'])
def cambiar_estado_presupuesto(id, nuevo_estado):
    if nuevo_estado not in ('pendiente', 'aprobado', 'rechazado'):
        flash('Estado inválido.', 'danger')
        return redirect(url_for('ver_presupuesto', id=id))
    db = get_db()
    execute(db, 'UPDATE presupuestos SET estado=? WHERE id=?', (nuevo_estado, id))
    db.commit(); db.close()
    flash(f'Presupuesto marcado como {nuevo_estado}.', 'success')
    return redirect(url_for('ver_presupuesto', id=id))

@app.route('/presupuestos/<int:id>/imprimir')
def imprimir_presupuesto(id):
    db = get_db()
    p     = fetchone(db, _presup_query(), (id,))
    items = fetchall(db, 'SELECT * FROM presupuesto_items WHERE presupuesto_id=? ORDER BY tipo, id', (id,))
    total = sum(i['cantidad'] * i['precio_unitario'] for i in items)
    db.close()
    return render_template('presupuesto_print.html', p=p, items=items, total=total, taller=TALLER)

@app.route('/presupuestos/<int:id>/eliminar', methods=['POST'])
def eliminar_presupuesto(id):
    db = get_db()
    execute(db, 'DELETE FROM presupuesto_items WHERE presupuesto_id=?', (id,))
    execute(db, 'DELETE FROM presupuestos WHERE id=?', (id,))
    db.commit(); db.close()
    flash('Presupuesto eliminado.', 'success')
    return redirect(url_for('presupuestos'))

# ─── EXPORTAR EXCEL ──────────────────────────────────────────────────────────

@app.route('/exportar/ordenes')
def exportar_ordenes():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Instala openpyxl para exportar.', 'danger')
        return redirect(url_for('ordenes'))
    db = get_db()
    rows = fetchall(db, '''
        SELECT o.id, o.fecha_ingreso, o.fecha_egreso, o.estado,
               c.nombre || ' ' || c.apellido AS cliente, c.telefono,
               v.marca, v.modelo, v.patente, v.anio, v.cilindrada, v.tipo_moto,
               o.diagnostico, o.descripcion_trabajo, o.repuestos
        FROM ordenes o
        JOIN clientes c ON o.cliente_id=c.id
        JOIN vehiculos v ON o.vehiculo_id=v.id
        ORDER BY o.fecha_ingreso DESC
    ''')
    db.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Órdenes'
    headers = ['N°','Ingreso','Egreso','Estado','Cliente','Teléfono',
               'Marca','Modelo','Dominio','Año','Cilindrada','Tipo','Diagnóstico','Trabajo','Repuestos']
    fill = PatternFill(start_color='1a2332', end_color='1a2332', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill; cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(list(row.values()))
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, download_name='ordenes_taller_motos.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/exportar/motos')
def exportar_motos():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Instala openpyxl para exportar.', 'danger')
        return redirect(url_for('motos'))
    db = get_db()
    rows = fetchall(db, '''
        SELECT v.id, v.marca, v.modelo, v.patente, v.anio, v.cilindrada, v.tipo_moto,
               v.color, v.motor, v.combustible,
               c.nombre || ' ' || c.apellido AS cliente
        FROM vehiculos v LEFT JOIN clientes c ON v.cliente_id=c.id
        ORDER BY v.marca
    ''')
    db.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Motos'
    headers = ['ID','Marca','Modelo','Dominio','Año','Cilindrada','Tipo','Color','Motor','Combustible','Cliente']
    fill = PatternFill(start_color='1a2332', end_color='1a2332', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill; cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(list(row.values()))
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, download_name='motos_taller.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/exportar/clientes')
def exportar_clientes():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Instala openpyxl para exportar.', 'danger')
        return redirect(url_for('clientes'))
    db = get_db()
    rows = fetchall(db, '''
        SELECT c.id, c.nombre, c.apellido, c.direccion, c.telefono,
               COUNT(v.id) AS motos
        FROM clientes c LEFT JOIN vehiculos v ON v.cliente_id=c.id
        GROUP BY c.id ORDER BY c.apellido
    ''')
    db.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Clientes'
    headers = ['ID','Nombre','Apellido','Dirección','Teléfono','Motos']
    fill = PatternFill(start_color='1a2332', end_color='1a2332', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill; cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(list(row.values()))
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, download_name='clientes_taller_motos.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
